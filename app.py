import streamlit as st
from openpyxl import Workbook
from io import BytesIO

PRONOUNS = ["yo","tú","vos","él/ella/Ud.","nosotros","ellos/ellas/Uds."]

SER = {
    "Indicative": {
        "Present": ["soy","eres","sos","es","somos","son"],
        "Preterite": ["fui","fuiste","fuiste","fue","fuimos","fueron"],
        "Imperfect": ["era","eras","eras","era","éramos","eran"],
        "Future": ["seré","serás","serás","será","seremos","serán"],
        "Conditional": ["sería","serías","serías","sería","seríamos","serían"]
    },
    "Progressive": {
        "Present": ["estoy siendo","estás siendo","estás siendo","está siendo","estamos siendo","están siendo"]
    },
    "Perfect": {
        "Present": ["he sido","has sido","has sido","ha sido","hemos sido","han sido"]
    },
    "Subjunctive": {
        "Present": ["sea","seas","seas","sea","seamos","sean"],
        "Imperfect": ["fuera","fueras","fueras","fuera","fuéramos","fueran"]
    },
    "Perfect Subjunctive": {
        "Present": ["haya sido","hayas sido","hayas sido","haya sido","hayamos sido","hayan sido"]
    },
    "Imperative": {
        "Affirmative": ["—","sé","sé","sea","seamos","sean"],
        "Negative": ["—","no seas","no seas","no sea","no seamos","no sean"]
    }
}

def to_excel(data, verb):
    wb = Workbook()
    ws = wb.active
    ws.title = verb

    row = 1

    for mood, tenses in data.items():
        ws.cell(row=row, column=1, value=mood)
        row += 1

        for tense, forms in tenses.items():
            ws.cell(row=row, column=1, value=tense)

            for i, p in enumerate(PRONOUNS):
                ws.cell(row=row, column=i+2, value=p)
                ws.cell(row=row+1, column=i+2, value=forms[i])

            row += 3

        row += 1

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

st.title("Spanish Verb Conjugator")

verb = st.text_input("Enter verb", "ser")

if st.button("Generate"):
    st.subheader(f"Conjugation: {verb}")

    for mood, tenses in SER.items():
        st.markdown(f"## {mood}")

        for tense, forms in tenses.items():
            st.markdown(f"**{tense}**")

            cols = st.columns(len(PRONOUNS))
            for i, col in enumerate(cols):
                col.write(PRONOUNS[i])
                col.write(forms[i])

    excel_file = to_excel(SER, verb)

    st.download_button(
        "Download Excel",
        excel_file,
        f"{verb}_conjugation.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
